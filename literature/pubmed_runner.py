import os
import math
import time
import pandas as pd
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
TOOL = "pubmed_automation"
EMAIL = os.getenv("NCBI_EMAIL", "sales@iziel.com")
API_KEY = os.getenv("NCBI_API_KEY")  # optional

DEFAULT_SLEEP = 0.34 if not API_KEY else 0.12  # request rate
MAX_RETRIES = 5

session = requests.Session()

# ----------------- Helpers -----------------
def _common_params():
    p = {"tool": TOOL, "email": EMAIL}
    if API_KEY:
        p["api_key"] = API_KEY
    return p

def safe_request(func, *args, **kwargs):
    retries = 0
    while True:
        try:
            return func(*args, **kwargs)
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 429 and retries < MAX_RETRIES:
                wait_time = (2 ** retries) * DEFAULT_SLEEP
                print(f"429 Too Many Requests. Retrying in {wait_time:.1f} seconds...")
                time.sleep(wait_time)
                retries += 1
                continue
            else:
                raise

def build_query(keyword: str, filters_csv: str, APPLY_ABSTRACT=True, APPLY_FREE=False, APPLY_FULL=False):
    parts = [f"({keyword})", "english[lang]", "humans[mh]"]
    avail_terms = []
    if APPLY_ABSTRACT:
        avail_terms.append("hasabstract[text]")
    if APPLY_FREE:
        avail_terms.append("free full text[sb]")
    if APPLY_FULL:
        avail_terms.append("full text[sb]")
    if avail_terms:
        parts.append("(" + " OR ".join(avail_terms) + ")")
    if filters_csv and filters_csv.strip():
        types = [t.strip() for t in filters_csv.split(",") if t.strip()]
        if types:
            parts.append("(" + " OR ".join([f'"{t}"[Publication Type]' for t in types]) + ")")
    return " AND ".join(parts)

def esearch_with_history(term: str, mindate=None, maxdate=None):
    time.sleep(0.4)
    params = {
        "db": "pubmed",
        "term": term,
        "retmode": "json",
        "usehistory": "y",
    }
    if mindate and maxdate:
        params.update({"datetype": "pdat", "mindate": mindate.replace("-", "/"), "maxdate": maxdate.replace("-", "/")})
    elif mindate:
        params.update({"datetype": "pdat", "mindate": mindate.replace("-", "/")})
    elif maxdate:
        params.update({"datetype": "pdat", "maxdate": maxdate.replace("-", "/")})

    params.update(_common_params())
    r = safe_request(session.get, BASE_URL + "esearch.fcgi", params=params, timeout=60)
    r.raise_for_status()
    data = r.json()
    res = data.get("esearchresult", {})
    return int(res.get("count", 0)), res.get("querykey"), res.get("webenv")

def safe_text(el):
    return (el.text or "").strip() if el is not None else ""

def xml_to_records(xml_text: str):
    root = ET.fromstring(xml_text)
    records = []
    for art in root.findall(".//PubmedArticle"):
        pmid = safe_text(art.find(".//PMID"))
        title = safe_text(art.find(".//ArticleTitle"))
        journal = safe_text(art.find(".//Journal/Title"))
        pubdate = safe_text(art.find(".//Journal/JournalIssue/PubDate/Year"))
        authors = ", ".join([
            f"{safe_text(a.find('LastName'))} {safe_text(a.find('Initials'))}".strip()
            for a in art.findall(".//AuthorList/Author")
        ])
        abstract = "\n".join([safe_text(a) for a in art.findall(".//AbstractText")])
        records.append({
            "PMID": pmid,
            "Title": title,
            "Journal": journal,
            "PubDate": pubdate,
            "Authors": authors,
            "Abstract": abstract,
            "PubMedURL": f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else ""
        })
    return records

def efetch_batch(qk: str, we: str, retstart: int, retmax: int = 200):
    params = {
        "db": "pubmed",
        "query_key": qk,
        "webenv": we,
        "retstart": retstart,
        "retmax": retmax,
        "retmode": "xml",
    }
    params.update(_common_params())
    r = safe_request(session.get, BASE_URL + "efetch.fcgi", params=params, timeout=120)
    r.raise_for_status()
    return r.text

# ----------------- CSV Merge -----------------
def merge_csvs(csv_dir: str, output_file: str):
    required_cols = ["PMID", "Title", "Journal", "PubDate", "Authors", "Abstract", "PubMedURL"]
    combined_rows = []

    for file_name in sorted(os.listdir(csv_dir)):
        if file_name.endswith(".csv") and file_name.startswith("#"):
            source = file_name.split("_")[0]
            df = pd.read_csv(os.path.join(csv_dir, file_name))
            for idx, row in df.iterrows():
                combined_rows.append({
                    "Source": source,
                    "KeywordNo": file_name.replace(".csv", ""),
                    "KeyCodeNo": f"{file_name.replace('.csv','')}.{idx+1}",
                    "PMID": row.get("PMID",""),
                    "Title": row.get("Title",""),
                    "Journal": row.get("Journal",""),
                    "PubDate": row.get("PubDate",""),
                    "Authors": row.get("Authors",""),
                    "Abstract": row.get("Abstract",""),
                    "PubMedURL": row.get("PubMedURL","")
                })

    combined_df = pd.DataFrame(combined_rows)
    combined_df.insert(0, "Sr.No", range(1, len(combined_df)+1))
    master_df = combined_df.drop_duplicates(subset="PMID", keep="first").reset_index(drop=True)
    master_df["Sr.No"] = range(1, len(master_df)+1)
    duplicates = combined_df.duplicated(subset="PMID", keep=False)
    highlight_df = combined_df.copy()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name="Combined", index=False)
        highlight_df.to_excel(writer, sheet_name="Duplicate", index=False)
        master_df.to_excel(writer, sheet_name="Master", index=False)

    wb = load_workbook(output_file)
    ws = wb["Duplicate"]
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row_idx, is_dup in enumerate(duplicates, start=2):
        if is_dup:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = fill
    wb.save(output_file)
    return output_file

# ----------------- MAIN PIPELINE -----------------
def run_pubmed_pipeline(keywords_path: str, output_dir: str, params: dict, project_id: str):

    df = pd.read_excel(keywords_path)
    df.columns = df.columns.str.strip()

    filtered_df = df[df["Keyword No."].astype(str).str.startswith("#")].copy()
    filters_col = filtered_df["Filters"] if "Filters" in filtered_df.columns else pd.Series([""] * len(filtered_df))
    keywords = list(zip(filtered_df["Keyword No."], filtered_df["Keywords"], filters_col.fillna("")))

    # ------------------------------------------------------------------
    # ✅ FIXED FOLDER STRUCTURE (NO NESTED PROJECT FOLDERS)
    # ------------------------------------------------------------------
    project_dir = output_dir                            # already: database/PRJ-1/literature
    csv_dir = os.path.join(project_dir, "All-CSV")

    os.makedirs(project_dir, exist_ok=True)
    os.makedirs(csv_dir, exist_ok=True)
    # ------------------------------------------------------------------

    for keyword_no, keyword, filters_csv in keywords:
        term = build_query(
            keyword, filters_csv,
            APPLY_ABSTRACT=params.get("abstract", True),
            APPLY_FREE=params.get("freeFullText", False),
            APPLY_FULL=params.get("fullText", False),
        )

        count, qk, we = esearch_with_history(
            term,
            params.get("fromDate"),
            params.get("toDate")
        )

        if count == 0:
            continue

        all_rows = []
        BATCH = 200
        loops = math.ceil(count / BATCH)

        for i in range(loops):
            retstart = i * BATCH
            time.sleep(DEFAULT_SLEEP)
            xml_text = efetch_batch(qk, we, retstart, BATCH)
            rows = xml_to_records(xml_text)
            all_rows.extend(rows)

        pd.DataFrame(all_rows).to_csv(
            os.path.join(csv_dir, f"{keyword_no}.csv"),
            index=False
        )

    merged_path = merge_csvs(
        csv_dir,
        os.path.join(project_dir, "All-Merged.xlsx")
    )

    return merged_path
